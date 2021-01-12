#!/usr/bin/env python3
# -*- coding: utf-8 -*-


"""
@author: Kosuke Asada
@date: 2021/01/12
@version: 0.1.9

Excel sheetにテーブル(pandas.DataFrame)を書式付きで書き込む.

@history:
    2020/11/30:
        初期版作成.
    
    2020/12/09:
        CopyExcelFormatに処理時間を表示.
    
    2020/12/14:
        色々追加.
        並列処理化.
    
    2020/12/16:
        debug_modeでシート名を表示.
    
    2020/12/18:
        copy_excel_format_from_temporary_filesに低スペックPC向けのsleepする秒数を追加.
    
    2020/12/23:
        package向けに修正.
"""



# --------------------------------------------------------------------------------
# Load modules
# --------------------------------------------------------------------------------

import sys, os
import gc
import shutil
import time
import copy
from multiprocess import Pool
from multiprocessing import Pool

from joblib import Parallel, delayed

import openpyxl as px
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter



# --------------------------------------------------------------------------------
# Functions
# --------------------------------------------------------------------------------


def pro_round(num, ndigits=0):
    """
    数字を四捨五入で丸める.

    Args:
        num: int or float
            丸めたい数字.

        ndigits: int, optional(default=0)
            丸めた後の小数部分の桁数.

    Returns:
        rounded: int or float
            丸めた後の数字.
    """
    num *= 10 ** ndigits
    rounded = ( 2 * num + 1 ) // 2
    rounded /= 10 ** ndigits

    if ndigits == 0:
        rounded = int(rounded)

    return rounded


def get_elapsed_time(start, print_time=True):
    """
    経過時間を取得する.
    使用前にtime.time()で処理開始時間を取得しておく.

    Args:
        start: float
            計測開始時間.
    
    Returns:
        elapsed_time: float
            経過時間.
    """
    end = time.time()
    elapsed_time = float(end - start)
    rounded_elapsed_time = pro_round(num=elapsed_time, ndigits=1)

    if print_time:
        print('elapsed time:', rounded_elapsed_time, 's')
    
    return elapsed_time


# from multiprocess import Pool
def multiprocess_func(func, args_list, processes=None, initializer=None, initargs=(), maxtasksperchild=None, chunksize=None):
    """
    マルチプロセスで並列処理を行う.
    
    Args:
        func: function
            並列処理したい関数.
        
        args_list: list
            関数の引数リスト.
        
        processes: int or None, optional(default=None)
            プロセス数.
    
    Returns:
        ret_list: list
            list of Returns
    
    Example:
        
        def func0(x, y):
            import time
            time.sleep(5)
            return [x+y, x-y]

        args_list0 = [
            [1, 1],
            [2, 2],
            [3, -1],
            [4, -2]
        ]

        ret0 = multiprocess_func(
            func = func0,
            args_list = args_list0,
            processes = 4
        )
        print('list of Returns:', ret0)
        # list of Returns: [[2, 0], [4, 0], [2, 4], [2, 6]]
    
    Remark:
        Multiprocessing example giving AttributeError
        (https://stackoverflow.com/questions/41385708/multiprocessing-example-giving-attributeerror)
        
        How to use multiprocessing pool.map with multiple arguments?
        (https://stackoverflow.com/questions/5442910/how-to-use-multiprocessing-pool-map-with-multiple-arguments)
    """
    with Pool(
        processes=processes,
        initializer=initializer,
        initargs=initargs,
        maxtasksperchild=maxtasksperchild
    ) as pool_obj:
        
        ret_list = pool_obj.starmap(func=func, iterable=args_list, chunksize=chunksize)
    
    return ret_list

mpf = multiprocess_func


# from joblib import Parallel, delayed
def parallel_func(func, args_list, args_type='list', n_jobs=-1, verbose=1, backend='loky', prefer='threads', batch_size='auto'):
    """
    execute function in parallel.
    joblibを使用. joblib内含め, multiprocessing moduleのエラーにより, multi　process非対応, multi threadsのみ対応.
    
    Args:
        func: function
            並列処理したい関数.
        
        args_list: list
            関数の引数リスト.
        
        args_type: str, optional(default='list')
            'list' or 'dict'
            引数の型.
        
        n_jobs: int, optional(default=-1)
            worker数.
        
        verbose: int, optional(default=0)
            min:0, max:50
            1以上の時, 途中経過が表示される.
        
        backend: str, optional(default='loky')
            'loky' or 'multiprocessing' or 'threading'
            'multiprocessing'はerrorが出やすい.
            並列処理の方法. マルチスレッドかマルチプロセス.
        
        prefer: str, optional(default='threads')
            'threads' or 'processes'
            並列処理の方法. マルチスレッドかマルチプロセス.
        
        batch_size: int or 'auto', optional(default='auto')
            バッチサイズ.
            同時に処理する処理数.
    
    Returns:
        ret_list: list
            list of Returns
    
    Examples:
        
        def func0(x, y):
            import time
            time.sleep(5)
            return [x+y, x-y]

        args_list0 = [
            [1, 1],
            [2, 2],
            [3, -1],
            [4, -2]
        ]

        args_list1 = [
            {'x': 1, 'y': 1},
            {'x': 2, 'y': 2},
            {'x': 3, 'y': -1},
            {'x': 4, 'y': -2}
        ]

        ret0 = parallel_func(func=func0, args_list=args_list0, args_type='list', verbose=1)
        print('type of args is list:', ret0)
        # type of args is list: [[2, 0], [4, 0], [2, 4], [2, 6]]

        ret1 = parallel_func(func=func0, args_list=args_list1, args_type='dict', verbose=2, prefer='processes')
        print('type of args is dict:', ret1)
        # type of args is dict: [[2, 0], [4, 0], [2, 4], [2, 6]]
    """
    ret_list = []
    
    if args_type=='list':
        
        ret_list = Parallel(
            n_jobs=n_jobs,
            verbose=verbose,
            backend=backend,
            prefer=prefer,
            batch_size=batch_size
        )([delayed(func)(*args) for args in args_list])
    
    elif args_type=='dict':
        ret_list = Parallel(
            n_jobs=n_jobs,
            verbose=verbose,
            backend=backend,
            prefer=prefer,
            batch_size=batch_size
        )([delayed(func)(**args) for args in args_list])
    
    return ret_list

plf = parallel_func


def pro_makedirs(dir_path):
    """
    ディレクトリを作成する.指定のディレクトリが存在しない場合のみ作成する.
    深い階層のディレクトリを指定した場合、途中階層のディレクトリも全て作成する.

    Arg:
        dir_path: str
            ディレクトリのパス.
    """
    dir_path = str(dir_path)
    if not os.path.isdir(dir_path):
        os.makedirs(dir_path)
    else:
        pass


def copy_excel_sheet_between_multibooks(input_excel_path, output_excel_path, sheet_name, copy_sheet_method='xlwings'):
    """
    excelファイル間を跨いでsheetをコピーする.
    
    Args:
        input_excel_path: str
            input excel path
        
        output_excel_path: str
            output excel path
        
        sheet_name: str
            sheet name to copy
        
        copy_sheet_method: str, optional(default='xlwings')
            'xlwings' or 'win32com'
    """

    # xlwingsを使用したシートのコピー. 別ブック対応.
    if copy_sheet_method == 'xlwings':

        import xlwings as xw

        input_wb = xw.Book(input_excel_path)
        output_wb = xw.Book(output_excel_path)

        input_ws = input_wb.sheets[sheet_name]
        input_ws.api.Copy(Before=output_wb.sheets(1).api)

        output_wb.save()
        output_wb.app.quit()


    # win32comを使用したシートのコピー. 別ブック対応.
    elif copy_sheet_method == 'win32com':

        from win32com.client import Dispatch

        xl_app = Dispatch('Excel.Application')
        # xl_app.Visible = True  # You can remove this line if you don't want the Excel application to be visible
        xl_app.Visible = False

        input_wb = xl_app.Workbooks.Open(Filename=input_excel_path)
        output_wb = xl_app.Workbooks.Open(Filename=output_excel_path)

        input_ws = input_wb.Worksheets(sheet_name)
        input_ws.Copy(Before=output_wb.Worksheets(sheet_name))

        output_wb.Close(SaveChanges=True)
        xl_app.Quit()

    else:
        pass


def output_temporary_excel(
    tmp_ceih,
    tmp_output_excel_dir_path,
    cef_manual_set_rows=None,
    cef_force_dimension_copy=False,
    cef_debug_mode=False,
    write_index=False,
    write_header=True,
    copy_values=False
    ):
    
    # 書式コピー元excelをopenpyxlで開く
    tmp_template_wb = px.load_workbook(tmp_ceih.template_excel_path)

    # 書式コピー元シート
    tmp_template_ws = tmp_template_wb[tmp_ceih.template_sheet_name]

    # 一時的に保存するworkbook
    tmp_output_wb = px.Workbook()

    # 一時的workbookにシートを作成
    tmp_output_wb.create_sheet(title=tmp_ceih.output_sheet_name)

    # CopyExcelFormatの作成
    cef = CopyExcelFormat(
        input_ws = tmp_template_ws,
        output_ws = tmp_output_wb[tmp_ceih.output_sheet_name],
        df = tmp_ceih.df,
        manual_set_rows = cef_manual_set_rows,
        force_dimension_copy = cef_force_dimension_copy,
        debug_mode = cef_debug_mode
    )

    cef.write_df2formatted_sheet(
        write_index = write_index,
        write_header = write_header,
        copy_values = copy_values
    )

    if cef_debug_mode:
        print('*' * 80)
        print('template_excel_path:{}'.format(str(tmp_ceih.template_excel_path)))
        print('template_sheet_name:{}'.format(str(tmp_ceih.template_sheet_name)))
        print('output_sheet_name:{}'.format(str(tmp_ceih.output_sheet_name)))
        print('*' * 80)
    
    del cef
    gc.collect()

    
    # デフォルトで作成される不要な1つ目のシートを削除する
    tmp_output_wb.remove(tmp_output_wb.worksheets[0])

    # 一時的excelの保存
    tmp_output_excel_path = tmp_output_excel_dir_path + tmp_ceih.output_sheet_name + '.xlsx'
    tmp_output_wb.save(tmp_output_excel_path)


def copy_excel_format(
    ceih_list,
    output_excel_path,
    cef_manual_set_rows=None,
    cef_force_dimension_copy=False,
    cef_debug_mode=False,
    write_index=False,
    write_header=True,
    copy_values=False
    ):
    """
    (並列処理を行わず,)excelファイルの書式のコピーとDataFrameの値書き込みを行う.

    Args:
        ceih_list: list of CopyExcelInfoHolder object
            list of CopyExcelInfoHolder object
        
        output_excel_path: str
            output excel file path
        
        cef_manual_set_rows: list of int[min_row, max_row] or None, optional(default=None)
            書式コピーを行う行数をマニュアルで指定する.

        cef_force_dimension_copy: bool, optional(default=False)
            強制的に行と列の幅などをコピー, 反映させる.

        cef_debug_mode: bool, optional(default=False)
            デバッグモード.

        write_index: bool, optional(default=False)
            indexを書き込むかどうか.

        write_header: bool, optional(default=False)
            headerを書き込むかどうか.

        copy_values: bool, optional(default=False)
            値のコピーを行うかどうか.
    """
    
    # 一旦最終出力先excelを作成
    output_wb = px.Workbook()

    for tmp_ceih in ceih_list:

        # 書式コピー元excelをopenpyxlで開く
        tmp_template_wb = px.load_workbook(tmp_ceih.template_excel_path)

        # 書式コピー元シート
        tmp_template_ws = tmp_template_wb[tmp_ceih.template_sheet_name]

        # workbookにシートを作成
        output_wb.create_sheet(title=tmp_ceih.output_sheet_name)

        # CopyExcelFormatの作成
        cef = CopyExcelFormat(
            input_ws = tmp_template_ws,
            output_ws = output_wb[tmp_ceih.output_sheet_name],
            df = tmp_ceih.df,
            manual_set_rows = cef_manual_set_rows,
            force_dimension_copy = cef_force_dimension_copy,
            debug_mode = cef_debug_mode
        )


        cef.write_df2formatted_sheet(
            write_index = write_index,
            write_header = write_header,
            copy_values = copy_values
        )

        if cef_debug_mode:
            print('*' * 80)
            print('template_excel_path:{}'.format(str(tmp_ceih.template_excel_path)))
            print('template_sheet_name:{}'.format(str(tmp_ceih.template_sheet_name)))
            print('output_sheet_name:{}'.format(str(tmp_ceih.output_sheet_name)))
            print('*' * 80)

        del cef
        gc.collect()


    # デフォルトで作成される不要な1つ目のシートを削除する
    output_wb.remove(output_wb.worksheets[0])

    # 保存
    output_wb.save(output_excel_path)


def output_temporary_excel_parallel(
    ceih_list,
    tmp_output_excel_dir_path='./tmp_output_excel/',
    parallel_method='joblib_multithreads',
    n_jobs=1,
    cef_manual_set_rows=None,
    cef_force_dimension_copy=False,
    cef_debug_mode=False,
    write_index=False,
    write_header=True,
    copy_values=False
    ):
    """
    並列処理を行い, 一時的な書式設定済みのexcelファイルを出力する.

    Args:
        ceih_list: list of CopyExcelInfoHolder object
            list of CopyExcelInfoHolder object
        
        tmp_output_excel_dir_path: str, optional(default='./tmp_output_excel/')
            temporary output excel file path
        
        parallel_method: str, optional(default='joblib_multithreads')
            'joblib_multithreads' or 'multiprocess'
            parallel method
        
        n_jobs: int, optional(default=1)
            number of workers

        cef_manual_set_rows: list of int[min_row, max_row] or None, optional(default=None)
            書式コピーを行う行数をマニュアルで指定する.

        cef_force_dimension_copy: bool, optional(default=False)
            強制的に行と列の幅などをコピー, 反映させる.

        cef_debug_mode: bool, optional(default=False)
            デバッグモード.

        write_index: bool, optional(default=False)
            indexを書き込むかどうか.

        write_header: bool, optional(default=False)
            headerを書き込むかどうか.

        copy_values: bool, optional(default=False)
            値のコピーを行うかどうか.
    """

    # 一時的なexcelを格納するディレクトリを作成
    pro_makedirs(tmp_output_excel_dir_path)

    """
    for tmp_ceih in ceih_list:
        output_temporary_excel(tmp_ceih, tmp_output_excel_dir_path)
    """

    if parallel_method == 'joblib_multithreads':

        # 並列処理
        parallel_func(
            func = output_temporary_excel,
            args_list = [
                [
                    tmp_ceih,
                    tmp_output_excel_dir_path,
                    cef_manual_set_rows,
                    cef_force_dimension_copy,
                    cef_debug_mode,
                    write_index,
                    write_header,
                    copy_values
                ] for tmp_ceih in ceih_list
            ],
            args_type = 'list',
            n_jobs = n_jobs
        )
    
    elif parallel_method == 'multiprocess':

        # 並列処理
        multiprocess_func(
            func = output_temporary_excel,
            args_list = [
                [
                    tmp_ceih,
                    tmp_output_excel_dir_path,
                    cef_manual_set_rows,
                    cef_force_dimension_copy,
                    cef_debug_mode,
                    write_index,
                    write_header,
                    copy_values
                ] for tmp_ceih in ceih_list
            ],
            processes = n_jobs
        )
    
    else:
        print('error: parallel_method is invalid.')


def copy_excel_format_from_temporary_files(
    ceih_list,
    output_excel_path,
    tmp_output_excel_dir_path='./tmp_output_excel/',
    copy_sheet_method='xlwings',
    sorted_sheet_names_list=None,
    del_tmp_dir=True,
    n_seconds_to_sleep=0,
    ):
    """
    一時的に出力した複数のexcelファイルをまとめて複数シートを持つ1つのexcelファイルとする.

    Args:
        ceih_list: list of CopyExcelInfoHolder object
            list of CopyExcelInfoHolder object
        
        output_excel_path: str
            output excel file path
        
        tmp_output_excel_dir_path: str, optional(default='./tmp_output_excel/')
            temporary output excel file path
        
        copy_sheet_method: str, optional(default='xlwings')
            'xlwings' or 'win32com'
        
        sorted_sheet_names_list: list of str
            sorted sheet names.
        
        del_tmp_dir: bool, optional(default=True)
            delete temporary directory.
        
        n_seconds_to_sleep: float, optional(default=True)
            Number of seconds to sleep for low spec PCs.
    """

    # 一旦最終出力先excelを作成
    output_wb = px.Workbook()

    # 保存
    output_wb.save(output_excel_path)

    # シート名リスト
    sheet_names_list = []

    # 低スペックPC向けのsleepする秒数
    # Number of seconds to sleep for low spec PCs.
    n_seconds_to_sleep = float(n_seconds_to_sleep)

    # 一時的に作成していたシートをコピーする
    for tmp_ceih in ceih_list:
        tmp_output_excel_path = tmp_output_excel_dir_path + tmp_ceih.output_sheet_name + '.xlsx'

        copy_excel_sheet_between_multibooks(
            input_excel_path = tmp_output_excel_path,
            output_excel_path = output_excel_path,
            sheet_name = tmp_ceih.output_sheet_name,
            copy_sheet_method = copy_sheet_method
        )

        sheet_names_list += [tmp_ceih.output_sheet_name]

        # 低スペックPC向けにtime.sleep()を挟む
        if n_seconds_to_sleep > 0:
            time.sleep(n_seconds_to_sleep)


    # シート名並び順を指定することも可能
    if sorted_sheet_names_list is None:
        sorted_sheet_names_list = sheet_names_list

    # 一時的excel格納ディレクトリを中身ごと削除
    if del_tmp_dir:
        shutil.rmtree(tmp_output_excel_dir_path)


    # シートの整理
    # シートを一旦ロードする
    output_wb = px.load_workbook(output_excel_path)

    # シートを並び替える
    n_sheets = len(output_wb.sheetnames)
    for tmp_sheet_name in sorted_sheet_names_list:
        output_wb.move_sheet(output_wb[tmp_sheet_name], offset=n_sheets)

    # デフォルトで作成される不要な1つ目のシートを削除する
    output_wb.remove(output_wb.worksheets[0])

    # 保存
    output_wb.save(output_excel_path)


def copy_excel_format_parallel(
    ceih_list,
    output_excel_path,
    tmp_output_excel_dir_path='./tmp_output_excel/',
    parallel_method='joblib_multithreads',
    n_jobs=1,
    copy_sheet_method='xlwings',
    sorted_sheet_names_list=None,
    del_tmp_dir=True,
    n_seconds_to_sleep=0,
    cef_manual_set_rows=None,
    cef_force_dimension_copy=False,
    cef_debug_mode=False,
    write_index=False,
    write_header=True,
    copy_values=False
    ):
    """
    並列処理を行い, 一時的な書式設定済みのexcelファイルを出力する.
    一時的に出力した複数のexcelファイルをまとめて複数シートを持つ1つのexcelファイルとする.

    Args:
        ceih_list: list of CopyExcelInfoHolder object
            list of CopyExcelInfoHolder object
        
        output_excel_path: str
            output excel file path
        
        tmp_output_excel_dir_path: str, optional(default='./tmp_output_excel/')
            temporary output excel file path
        
        parallel_method: str, optional(default='joblib_multithreads')
            'joblib_multithreads' or 'multiprocess'
            parallel method
        
        n_jobs: int, optional(default=1)
            number of workers
        
        copy_sheet_method: str, optional(default='xlwings')
            xlwings or win32com
        
        sorted_sheet_names_list: list of str
            sorted sheet names
        
        del_tmp_dir: bool, optional(default=True)
            delete temporary directory
        
        n_seconds_to_sleep: float, optional(default=True)
            Number of seconds to sleep for low spec PCs.
        
        cef_manual_set_rows: list of int[min_row, max_row] or None, optional(default=None)
            書式コピーを行う行数をマニュアルで指定する.

        cef_force_dimension_copy: bool, optional(default=False)
            強制的に行と列の幅などをコピー, 反映させる.

        cef_debug_mode: bool, optional(default=False)
            デバッグモード.
        
        write_index: bool, optional(default=False)
            indexを書き込むかどうか.
        
        write_header: bool, optional(default=False)
            headerを書き込むかどうか.
        
        copy_values: bool, optional(default=False)
            値のコピーを行うかどうか.
    """

    # 並列処理を行い, 一時的な書式設定済みのexcelファイルを出力する.
    output_temporary_excel_parallel(
        ceih_list = ceih_list,
        tmp_output_excel_dir_path = tmp_output_excel_dir_path,
        parallel_method = parallel_method,
        n_jobs = n_jobs,
        cef_manual_set_rows = cef_manual_set_rows,
        cef_force_dimension_copy = cef_force_dimension_copy,
        cef_debug_mode = cef_debug_mode,
        write_index = write_index,
        write_header = write_header,
        copy_values = copy_values
    )
    
    # 一時的に出力した複数のexcelファイルをまとめて複数シートを持つ1つのexcelファイルとする.
    copy_excel_format_from_temporary_files(
        ceih_list = ceih_list,
        output_excel_path = output_excel_path,
        tmp_output_excel_dir_path = tmp_output_excel_dir_path,
        copy_sheet_method = copy_sheet_method,
        sorted_sheet_names_list = sorted_sheet_names_list,
        del_tmp_dir = del_tmp_dir,
        n_seconds_to_sleep = n_seconds_to_sleep
    )



# --------------------------------------------------------------------------------
# Classes
# --------------------------------------------------------------------------------


# CopyExcelInfoHolder
class CopyExcelInfoHolder():
    def __init__(
        self,
        template_excel_path,
        template_sheet_name,
        output_sheet_name,
        df
    ):  
        self.template_excel_path = template_excel_path
        self.template_sheet_name = template_sheet_name
        self.output_sheet_name = output_sheet_name
        self.df = df


# CopyExcelFormat
class CopyExcelFormat():
    """
    Excel sheetにpandas.DataFrameを書式付きで書き込む.

    Attributes:
        input_ws: openpyxl.worksheet.worksheet.Worksheet
            書式コピー元worksheetオブジェクト.
        
        output_ws: openpyxl.worksheet.worksheet.Worksheet
            コピー先worksheetオブジェクト.
        
        df: pandas.DataFrame
            書き込むDataFrame.
        
        manual_set_rows: list of int[min_row, max_row] or None, optional(default=None)
            書式コピーを行う行数をマニュアルで指定する.

        force_dimension_copy: bool, optional(default=False)
            強制的に行と列の幅などをコピー, 反映させる.

        debug_mode: bool, optional(default=False)
            デバッグモード.
    
    Example:
        # CopyExcelFormatインスタンスの作成
        cef = CopyExcelFormat(
            input_ws = input_ws, # 書式コピー元wsオブジェクト(要修正)
            output_ws = output_ws,
            df = df,
            manual_set_rows = None,
            force_dimension_copy = False,
            debug_mode = False
        )

        # 書式書き込みの実行
        cef.write_df2formatted_sheet(
            write_index = False,
            write_header = False,
            copy_values = False
        )
    """

    def __init__(
        self,
        input_ws=None,
        output_ws=None,
        df=None,
        manual_set_rows=None,
        force_dimension_copy=False,
        debug_mode=False
        ):
        """
        CopyExcelFormatクラスのコンストラクタ.
        """

        self.set_input_ws(input_ws)
        self.set_output_ws(output_ws)
        self.set_df(df)
        self.manual_set_rows = manual_set_rows
        self.force_dimension_copy = force_dimension_copy
        self.debug_mode = debug_mode

        if self.debug_mode:
            self.init_time = time.time()
    

    def set_input_ws(self, input_ws=None):
        self.input_ws = input_ws


    def set_output_ws(self, output_ws=None):
        self.output_ws = output_ws
    

    def set_df(self, df=None):
        """
        pandas.DataFrameをセットする

        Args:
            df: pandas.DataFrame        
        """

        self.df = df
        
        if df is None:
            self.df_shape_list = None
            self.df_n_rows = None
            self.df_n_cols = None
        
        else:
            self.df_shape_list = list(df.shape)
            self.df_n_rows = self.df_shape_list[0]
            self.df_n_cols = self.df_shape_list[1]


    # private method
    def __copy_sheet_format(self, copy_values=False):
        """
        Excel sheetの書式を条件付き含めてコピーし, 別のsheetに反映させる.
        
        Args:            
            copy_values: bool, optional(default=False)
                値のコピーを行うかどうか.
        """
        
        if self.manual_set_rows is None:

            # 書式のコピー
            for row in self.input_ws.rows:
                for input_cell in row:

                    # 値コピーする場合
                    if copy_values:
                        output_cell = self.output_ws.cell(
                            row=input_cell.row,
                            column=input_cell.column,
                            value=input_cell.value
                        )
                    
                    # 値コピーしない場合
                    else:
                        output_cell = self.output_ws.cell(
                            row=input_cell.row,
                            column=input_cell.column
                        )

                    # cellにスタイルが指定されていた場合
                    if input_cell.has_style:
                        output_cell.font = copy.copy(input_cell.font)
                        output_cell.border = copy.copy(input_cell.border)
                        output_cell.fill = copy.copy(input_cell.fill)
                        output_cell.number_format = copy.copy(input_cell.number_format)
                        output_cell.protection = copy.copy(input_cell.protection)
                        output_cell.alignment = copy.copy(input_cell.alignment)

                    del input_cell, output_cell
                    gc.collect()
        
        elif type(self.manual_set_rows) == list:

            # 書式のコピー
            # manual_set_rows = [min_row, max_row]
            for row in self.input_ws.iter_rows(min_row=self.manual_set_rows[0], max_row=self.manual_set_rows[1]):
                for input_cell in row:

                    # 値コピーする場合
                    if copy_values:
                        output_cell = self.output_ws.cell(
                            row=input_cell.row,
                            column=input_cell.column,
                            value=input_cell.value
                        )
                    
                    # 値コピーしない場合
                    else:
                        output_cell = self.output_ws.cell(
                            row=input_cell.row,
                            column=input_cell.column
                        )

                    # cellにスタイルが指定されていた場合
                    if input_cell.has_style:
                        output_cell.font = copy.copy(input_cell.font)
                        output_cell.border = copy.copy(input_cell.border)
                        output_cell.fill = copy.copy(input_cell.fill)
                        output_cell.number_format = copy.copy(input_cell.number_format)
                        output_cell.protection = copy.copy(input_cell.protection)
                        output_cell.alignment = copy.copy(input_cell.alignment)

                    del input_cell, output_cell
                    gc.collect()

        else:
            print('error: manual set rows type error.')
        

        # 処理時間表示
        if self.debug_mode:
            print('to copy cell format end.')
            get_elapsed_time(self.wd2fs_start_time, print_time=True)
            print()


        # 条件付き書式のコピー
        self.output_ws.conditional_formatting = copy.copy(self.input_ws.conditional_formatting)

        # その他諸々のコピー
        # self.output_ws.conditional_formatting = self.input_ws.conditional_formatting
        self.output_ws.page_margins = copy.copy(self.input_ws.page_margins)
        self.output_ws.page_setup = copy.copy(self.input_ws.page_setup)
        self.output_ws.col_breaks = copy.copy(self.input_ws.col_breaks)
        self.output_ws.row_breaks = copy.copy(self.input_ws.row_breaks)
        self.output_ws.sheet_format = copy.copy(self.input_ws.sheet_format)
        self.output_ws.sheet_properties = copy.copy(self.input_ws.sheet_properties)
        self.output_ws._print_area = copy.copy(self.input_ws._print_area)
        self.output_ws.views = copy.copy(self.input_ws.views)

        # セル結合状態をコピー, 反映させる.
        for cell_range in self.input_ws.merged_cells.ranges:
            self.output_ws.merge_cells(range_string=cell_range.coord)

        # 強制的に行と列の幅をコピー, 反映させる.
        if self.force_dimension_copy:

            # 列の幅と表示非表示をコピー
            for col_idx in range(self.input_ws.max_column):
                col_name = get_column_letter(col_idx + 1)

                # 列の幅
                self.output_ws.column_dimensions[col_name].width = copy.copy(self.input_ws.column_dimensions[col_name].width)

                # 表示非表示
                self.output_ws.column_dimensions[col_name].hidden = copy.copy(self.input_ws.column_dimensions[col_name].hidden)

            # 行の高さと表示非表示をコピー
            for row_idx in range(self.input_ws.max_row):
                row_name = row_idx + 1

                # 行の高さ
                self.output_ws.row_dimensions[row_name].height = copy.copy(self.input_ws.row_dimensions[row_name].height)

                # 表示非表示
                self.output_ws.row_dimensions[row_name].hidden = copy.copy(self.input_ws.row_dimensions[row_name].hidden)
        
        else:

            # 列の幅と表示非表示をコピー
            for col_name, col_obj in self.input_ws.column_dimensions.items():

                # 列の幅
                self.output_ws.column_dimensions[col_name].width = copy.copy(col_obj.width)

                # 表示非表示
                self.output_ws.column_dimensions[col_name].hidden = copy.copy(col_obj.hidden)

            # 行の高さと表示非表示をコピー
            for row_name, row_obj in self.input_ws.row_dimensions.items():

                # 行の高さ
                self.output_ws.row_dimensions[row_name].height = copy.copy(row_obj.height)

                # 表示非表示
                self.output_ws.row_dimensions[row_name].hidden = copy.copy(row_obj.hidden)

    
    # private method
    def __write_df2sheet(self, write_index=False, write_header=False):
        """
        Excel sheetにpandas.DataFrameを書き込む.
        
        Args:
            write_index: bool, optional(default=False)
                indexを書き込むかどうか.
            
            write_header: bool, optional(default=False)
                headerを書き込むかどうか.
        """
        
        # DataFrameを1行ずつ書き込む.
        for df_row in dataframe_to_rows(self.df, index=write_index, header=write_header):
            self.output_ws.append(df_row)
    

    def write_df2formatted_sheet(self, write_index=False, write_header=True, copy_values=False):
        """
        Excel sheetにpandas.DataFrameを書式付きで書き込む.

        Args:            
            write_index: bool, optional(default=False)
                indexを書き込むかどうか.
            
            write_header: bool, optional(default=False)
                headerを書き込むかどうか.
            
            copy_values: bool, optional(default=False)
                値のコピーを行うかどうか.
        """

        if self.debug_mode:
            self.wd2fs_start_time = time.time()

            print('*' * 80)
            print('sheet name: {}'.format(str(self.output_ws.title)))
            print()

        # DataFrameの書き込み
        try:

            # DataFrameの書き込み
            self.__write_df2sheet(
                write_index = write_index,
                write_header = write_header
            )

        except Exception as e:
            print('error: write df to sheet.')
            print(e)

        # 処理時間表示
        if self.debug_mode:
            print('to write df to sheet end.')
            get_elapsed_time(self.wd2fs_start_time, print_time=True)
            print()

        # 書式のコピー        
        try:
            
            # 書式のコピー
            self.__copy_sheet_format(
                copy_values = copy_values
            )

        except Exception as e:
            print('error: copy sheet format.')
            print(e)
        
        # 処理時間表示
        if self.debug_mode:
            print('to copy format end.')
            get_elapsed_time(self.wd2fs_start_time, print_time=True)
            print()
            print('*' * 80)
            print()



    



